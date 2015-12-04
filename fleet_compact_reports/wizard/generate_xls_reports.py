# -*- coding: utf-8 -*-

import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.styles import Font, Color
from openpyxl.styles import colors
from openpyxl import Workbook
import base64
import cStringIO
from openpyxl.worksheet.dimensions import RowDimension, ColumnDimension
from openerp import models, fields, api
from datetime import datetime

MONTHS = [(1, 'Januar'),
          (2, 'Februar'),
          (3, 'Mart'),
          (4, 'April'),
          (5, 'Maj'),
          (6, 'Jun'),
          (7, 'Jul'),
          (8, 'Avgust'),
          (9, 'Septembar'),
          (10, 'Octobar'),
          (11, 'Novembar'),
          (12, 'Decembar')]

YEARS = [(2015, '2015'),
         (2016, '2016'),
         (2017, '2017'),
         (2018, '2018'),
         (2019, '2019'),
         (2020, '2020')]

REPORT_TYPES = [
    ('obracun_plate', 'Obračun plate'),
    ('pregled_transporta_za_vozaca', 'Pregled transporta za vozača')
]


class FleetExcelReports(models.TransientModel):
    _name = 'fleet.compact.excel.report.selection'
    _description = 'Fleet excel report selection'

    @api.multi
    def _reopen(self):
        return {'type': 'ir.actions.act_window',
                'name': 'Preuzimanje izvještaja',
                'view_mode': 'form',
                'view_type': 'form',
                'res_id': self.id,
                'res_model': 'fleet.compact.excel.report.selection',
                'target': 'new',
                }

    @api.multi
    def _get_obracun_plate(self):
        wb = openpyxl.Workbook(encoding='utf-8')
        ws = wb.create_sheet(0, self.employee_id.name)

        for i in range(1, 30):
            for j in range(1, 10):
                ws.cell(row=i, column=j).alignment.vertical = "center"
                ws.cell(row=i, column=j).alignment.horizontal = "center"

        ws['D5'] = 'ZARADA ZA MJESEC ' + dict(MONTHS).get(self.month).upper() + ' ' + dict(YEARS).get(
            self.year).upper() + '. GODINE'
        ws['D5'].font = Font(bold=True)

        ws['B7'] = "Vozac: " + self.employee_id.name
        ws['B7'].font = Font(bold=True)
        # ws['B7'].style.alignment = Alignment(horizontal="center", vertical="center", wrapText=True)

        ws['B8'] = 'R. br.'
        ws['B8'].font = Font(bold=True)
        ws['B8'].border = Border(left=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))
        # ws['B8'].style.alignment = Alignment(horizontal="center", vertical="center", wrapText=True)

        ws['C8'] = 'Br. koverte'
        ws['C8'].font = Font(bold=True)
        ws['C8'].border = Border(bottom=Side(style='thick'), top=Side(style='thick'))
        # ws['C8'].style.alignment = Alignment(horizontal="center", vertical="center", wrapText=True)

        ws['D8'] = 'Datum'
        ws['D8'].font = Font(bold=True)
        ws['D8'].border = Border(bottom=Side(style='thick'), top=Side(style='thick'))
        # ws['D8'].style.alignment = Alignment(horizontal="center", vertical="center", wrapText=True)

        ws['E8'] = 'Br. fakture'
        ws['E8'].font = Font(bold=True)
        ws['E8'].border = Border(top=Side(style='thick'))
        # ws['E8'].style.alignment = Alignment(horizontal="center", vertical="center", wrapText=True)

        ws['F8'] = 'Relacija'
        ws['F8'].font = Font(bold=True)
        ws['F8'].border = Border(top=Side(style='thick'))
        # ws['F8'].style.alignment = Alignment(horizontal="center", vertical="center", wrapText=True)

        ws['G8'] = 'Količina\n(tona)'
        ws['G8'].font = Font(bold=True)
        # ws['G8'].style.alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
        ws['G8'].border = Border(top=Side(style='thick'))

        ws['H8'] = 'Cijena\n€/Toni'
        ws['H8'].font = Font(bold=True)
        ws['H8'].border = Border(top=Side(style='thick'))
        # ws['H8'].style.alignment = Alignment(horizontal="center", vertical="center", wrapText=True)

        ws['I8'] = 'Vrijednost\n€'
        ws['I8'].font = Font(bold=True)
        ws['I8'].border = Border(top=Side(style='thick'))
        # ws['I8'].style.alignment = Alignment(horizontal="center", vertical="center", wrapText=True)

        ws['J8'] = 'Vrjednost\nKM'
        ws['J8'].font = Font(bold=True)
        ws['J8'].border = Border(top=Side(style='thick'))
        # ws['J8'].style.alignment = Alignment(horizontal="center", vertical="center", wrapText=True)

        ws['K8'] = 'Zarada vozača'
        ws['K8'].font = Font(bold=True)
        ws['K8'].border = Border(right=Side(style='thick'), top=Side(style='thick'))
        # ws['K8'].style.alignment = Alignment(horizontal="center", vertical="center", wrapText=True)

        ws.row_dimensions[8] = RowDimension(worksheet=ws, height=40)
        ws.column_dimensions['F'] = ColumnDimension(worksheet=ws, width=40)
        ws.column_dimensions['I'] = ColumnDimension(worksheet=ws, width=12)
        ws.column_dimensions['J'] = ColumnDimension(worksheet=ws, width=12)

        buf = cStringIO.StringIO()
        wb.save(buf)
        buf.seek(0)
        out = base64.encodestring(buf.read())
        buf.close()

        self.data = out
        self.name = "obracun_plate.xlsx"  # +self.employee_id.name+"_"+dict(MONTHS).get(self.month)+"_"+dict(YEARS).get(self.year)
        self.state = 'get'
        return True

    @api.multi
    def _get_pregled_transporta_za_vozaca(self):
        # bitno je povaditi transporte u odredjenom vremenskom periodu
        wb = openpyxl.Workbook()
        ws = wb.create_sheet(0, self.employee_id.name)

        buf = cStringIO.StringIO()
        wb.save(buf)
        buf.seek(0)
        out = base64.encodestring(buf.read())
        buf.close()

        self.data = out
        self.name = "pregled_transporta"
        self.state = 'get'
        return True

    name = fields.Char('File name', readonly=True)
    data = fields.Binary('File', readonly=True)
    report_type = fields.Selection(REPORT_TYPES, 'Odaberite željeni izvještaj', required=True, default='obracun_plate')
    employee_id = fields.Many2one('hr.employee', 'Odaberite zaposlenog')
    date_start = fields.Date('Početni datum')
    date_stop = fields.Date('Krajnji datum')
    month = fields.Selection(MONTHS, 'Odaberite mjesec')
    year = fields.Selection(YEARS, 'Odaberite godinu')
    state = fields.Selection([('choose', 'choose'), ('get', 'get')], default='choose')

    @api.multi
    def get_report(self):
        if self.report_type == 'obracun_plate':
            self._get_obracun_plate()
        if self.report_type == 'pregled_transporta_za_vozaca':
            self._get_pregled_transporta_za_vozaca()
        return {'type': 'ir.actions.act_window',
                'name': 'Izbor izvještaja',
                'view_mode': 'form',
                'view_type': 'form',
                'res_id': self.id,
                'res_model': 'fleet.compact.excel.report.selection',
                'target': 'new',
                }

# vim:expandtab:smartindent:tabstop=4:softtabstop=4:shiftwidth=4:
